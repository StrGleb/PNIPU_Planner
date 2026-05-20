import re
import uuid
from dataclasses import dataclass
from datetime import datetime, time
from typing import List, Tuple, Set
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


# ====== CONSTS ======
SHEET_NAME = "Лист1"
TIMETABLE_CELL_NAME = "A1"

LESSON_DURATION_MINUTES = 90
LESSONS_PER_DAY = 12
UNKNOWN_VALUE = "Unknown"

FIRST_SUBGROUP_COL = 3
SECOND_SUBGROUP_COL = 4
FIRST_LESSON_ROW = 4
LAST_LESSON_ROW = 75
SUBGROUP_ROW = 3
TIME_COLUMN_INDEX = 2


REGEX_TIME_NAME_TYPE = re.compile(r"([0-9]{1,2}:[0-9]{1,2})* *(.+?) (\(лек\)|\(пр\)|\(лаб\)|\(кср\)) (.*)")
REGEX_TEACHER_ROLE = re.compile(
    r"(асс\.)|(доц\.)|(зав\.)|(куратор)|(научный руководитель)|(ст\. *пр\.)|(пр\.)|(преп\.)|(проф\.)|(профессор)|(тренер-преподаватель)"
)
REGEX_TEACHER_LOCATION = re.compile(r"(.*?(?:[А-Я]\.)+) *(.*)")


@dataclass
class LessonInsertParams:
    staging_id: str
    subject: str
    category: str
    day: int
    time_start: int
    time_end: int
    repeat_rule: int
    timetable: str

@dataclass
class SubgroupAssignment:
    staging_id: str
    subgroup: str

@dataclass
class TeacherLocationAssignmentInsert:
    staging_id: str
    teacher: str
    location: str

@dataclass
class Lesson:
    raw_name: str
    insert_params: LessonInsertParams

@dataclass
class TeacherLocationAssignment:
    teacher: str
    location: str

    def __post_init__(self):
        self.teacher = self.teacher.strip()
        self.location = self.location.strip()

class PendingData:
    def __init__(self):
        self.subgroups_assignments: List[SubgroupAssignment] = []
        self.teacher_location_assignments: List[TeacherLocationAssignmentInsert] = []

def normalize_spaces(s: str) -> str:
    return " ".join(s.split())

def remove_all_spaces(s: str) -> str:
    return "".join(ch for ch in s if ch not in " \t\n\r")

def parse_time_hhmm(value: str) -> time:
    return datetime.strptime(value.strip(), "%H:%M").time()

def time_to_minutes(t: time) -> int:
    return t.hour * 60 + t.minute

def get_timetable_title(ws: Worksheet) -> str:
    val = ws[TIMETABLE_CELL_NAME].value
    if val is None:
        raise ValueError("Timetable title not found in A1")
    return str(val)

def get_subgroup_name(ws: Worksheet, col_index: int) -> str:
    val = ws.cell(row = SUBGROUP_ROW, column = col_index).value
    if val is None:
        raise ValueError(f"Subgroup name not found at col={col_index}")

    subgroup = remove_all_spaces(str(val))

    if "пг" not in subgroup:
        subgroup_number = col_index - 2
        subgroup += f"({subgroup_number}пг)"

    return subgroup


def get_lesson_start_time(ws: Worksheet, row_index: int) -> time:
    time_row = row_index if row_index % 2 == 0 else row_index - 1

    for r in range(time_row, 0, -2):
        val = ws.cell(row = r, column = TIME_COLUMN_INDEX).value

        if val is None or str(val).strip() == "":
            continue

        if isinstance(val, datetime):
            return val.time()

        return parse_time_hhmm(str(val))

    raise ValueError(f"Time value not found above row={row_index}")

def get_cell_merge_height(ws: Worksheet, col_index: int, row_index: int) -> int:
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        if min_col <= col_index <= max_col and min_row <= row_index <= max_row:
            return max_row - min_row + 1
    return 1

def determine_repeat_rule(ws: Worksheet, col_index: int, row_index: int) -> Tuple[int, bool]:
    merge_height = get_cell_merge_height(ws, col_index, row_index)

    if merge_height == 2:
        if row_index % 2 != 0:
            return 0, True
        return 0, False

    return (row_index % 2) + 1, False

def extract_teacher_and_location(segment: str) -> Tuple[str, str]:
    matches = REGEX_TEACHER_LOCATION.search(segment)
    if not matches:
        return UNKNOWN_VALUE, UNKNOWN_VALUE

    teacher = matches.group(1).strip() or UNKNOWN_VALUE
    location = matches.group(2).strip() or UNKNOWN_VALUE

    return teacher, location

def parse_teacher_locations(teachers_location_string: str) -> List[TeacherLocationAssignment]:
    role_matches = list(REGEX_TEACHER_ROLE.finditer(teachers_location_string))

    if not role_matches:
        return [TeacherLocationAssignment(UNKNOWN_VALUE, UNKNOWN_VALUE)]

    assignments: List[TeacherLocationAssignment] = []

    for i, match in enumerate(role_matches):
        start_idx = match.start()
        end_idx = len(teachers_location_string)

        if i < len(role_matches) - 1:
            end_idx = role_matches[i + 1].start()

        segment = teachers_location_string[start_idx:end_idx]
        teacher, location = extract_teacher_and_location(segment)
        assignments.append(TeacherLocationAssignment(teacher, location))

    assignments.sort(key = lambda x: (x.teacher, x.location))
    return assignments



class Parser:
    def __init__(self):
        self.subgroups: Set[str] = set()
        self.teachers: Set[str] = set()
        self.locations: Set[str] = set()
        self.subjects: Set[str] = set()
        self.timetables: Set[str] = set()

        self.lessons: List[Lesson] = []
        self.subgroups_assignments: List[SubgroupAssignment] = []
        self.teacher_location_assignments: List[TeacherLocationAssignmentInsert] = []

    def parse_lessons_from_bytes(self, file_bytes: bytes):
        wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)

        if SHEET_NAME not in wb.sheetnames:
            raise ValueError(f"Sheet '{SHEET_NAME}' not found")

        ws = wb[SHEET_NAME]
        timetable_name = get_timetable_title(ws)

        all_lessons: List[Lesson] = []

        for col_index in range(FIRST_SUBGROUP_COL, SECOND_SUBGROUP_COL + 1):
            try:
                subgroup = get_subgroup_name(ws, col_index)
            except ValueError:
                continue

            pending = PendingData()
            lessons = self._parse_subgroup_lessons(ws, col_index, subgroup, timetable_name, pending)

            if not self._is_duplicate_subgroup(all_lessons, lessons):
                all_lessons.extend(lessons)

                self.subgroups.add(subgroup)
                self.subgroups_assignments.extend(pending.subgroups_assignments)
                self.teacher_location_assignments.extend(pending.teacher_location_assignments)

        self.lessons.extend(all_lessons)
        self.timetables.add(timetable_name)

    def _is_duplicate_subgroup(self, existing: List[Lesson], new: List[Lesson]) -> bool:
        return [x.raw_name for x in existing] == [x.raw_name for x in new]

    def _parse_subgroup_lessons(
        self,
        ws: Worksheet,
        col_index: int,
        subgroup_name: str,
        timetable_name: str,
        pending: PendingData
    ) -> List[Lesson]:

        day = 0
        lessons: List[Lesson] = []

        for row_index in range(FIRST_LESSON_ROW, LAST_LESSON_ROW + 1):
            if (row_index - FIRST_LESSON_ROW) % LESSONS_PER_DAY == 0:
                day += 1

            lessons.extend(
                self._parse_lesson_cell(ws, col_index, row_index, subgroup_name, timetable_name, day, pending)
            )

        return lessons

    def _parse_lesson_cell(
        self,
        ws: Worksheet,
        col_index: int,
        row_index: int,
        subgroup_name: str,
        timetable_name: str,
        day: int,
        pending: PendingData
    ) -> List[Lesson]:

        raw_value = ws.cell(row=row_index, column=col_index).value

        if raw_value is None or str(raw_value).strip() == "":
            return []

        raw_value = normalize_spaces(str(raw_value))

        repeat_rule, should_skip = determine_repeat_rule(ws, col_index, row_index)
        if should_skip:
            return []

        try:
            time_start = get_lesson_start_time(ws, row_index)
        except ValueError:
            return []

        lesson_names = raw_value.split(" / ")
        lessons: List[Lesson] = []

        for lesson_name in lesson_names:
            lessons.append(
                self._parse_lesson(
                    lesson_name,
                    subgroup_name,
                    timetable_name,
                    day,
                    time_start,
                    repeat_rule,
                    pending
                )
            )

        return lessons

    def _parse_lesson(
        self,
        raw_name: str,
        subgroup_name: str,
        timetable_name: str,
        day: int,
        default_start_time: time,
        repeat_rule: int,
        pending: PendingData
    ) -> Lesson:

        staging_id = str(uuid.uuid4())

        insert_params = LessonInsertParams(
            staging_id=staging_id,
            subject=raw_name,
            category=UNKNOWN_VALUE,
            day=day,
            time_start=time_to_minutes(default_start_time),
            time_end=time_to_minutes(default_start_time) + LESSON_DURATION_MINUTES,
            repeat_rule=repeat_rule,
            timetable=timetable_name
        )

        pending.subgroups_assignments.append(
            SubgroupAssignment(staging_id=staging_id, subgroup=subgroup_name)
        )

        matches = REGEX_TIME_NAME_TYPE.match(raw_name)

        if not matches:
            self.subjects.add(raw_name)
            self.locations.add(UNKNOWN_VALUE)
            self.teachers.add(UNKNOWN_VALUE)

            pending.teacher_location_assignments.append(
                TeacherLocationAssignmentInsert(
                    staging_id=staging_id,
                    teacher=UNKNOWN_VALUE,
                    location=UNKNOWN_VALUE
                )
            )

            return Lesson(raw_name=raw_name, insert_params=insert_params)

        start_time = default_start_time
        if matches.group(1):
            start_time = parse_time_hhmm(matches.group(1))

        subject = matches.group(2)
        category = matches.group(3)
        teacher_location_str = matches.group(4)

        insert_params.subject = subject
        insert_params.category = category
        insert_params.time_start = time_to_minutes(start_time)
        insert_params.time_end = time_to_minutes(start_time) + LESSON_DURATION_MINUTES

        self.subjects.add(subject)

        for assignment in parse_teacher_locations(teacher_location_str):
            self.teachers.add(assignment.teacher)
            self.locations.add(assignment.location)

            pending.teacher_location_assignments.append(
                TeacherLocationAssignmentInsert(
                    staging_id=staging_id,
                    teacher=assignment.teacher,
                    location=assignment.location
                )
            )

        return Lesson(raw_name=raw_name, insert_params=insert_params)
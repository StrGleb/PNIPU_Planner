from io import BytesIO
from datetime import datetime
import re
from openpyxl import load_workbook


WEEKDAY_MAP = {
    "Понедельник": 1,
    "Вторник": 2,
    "Среда": 3,
    "Четверг": 4,
    "Пятница": 5,
    "Суббота": 6,
    "Воскресенье": 7,
}


def parse_date(value):
    if not value:
        return ""

    s = str(value).strip()
    if len(s) > 5:
        return s

    return f"{s}.2026"


def parse_session_line(raw: str):
    raw = str(raw).replace("\n", " ").replace("\r", " ").strip()

    time_start = raw[:5]
    rest = raw[6:].strip()

    if "Консультация" in rest:
        room = rest.replace("Консультация", "").strip()

        return {
            "time_start": time_start,
            "time_end": "",
            "subject": "Консультация",
            "lesson_type": "консультация",
            "teacher": "",
            "room": room
        }


    room_match = re.search(r"\d.*$", rest)
    room = room_match.group(0).strip() if room_match else ""

    if room_match:
        main_part = rest[:room_match.start()].strip()
    else:
        main_part = rest


    teacher_match = re.search(
        r"(доц\.|асс\.|преп\.|ст\.|проф\.|куратор|научный руководитель).*",
        main_part
    )

    if teacher_match:
        teacher = teacher_match.group(0).strip()
        subject = main_part[:teacher_match.start()].strip()
    else:
        teacher = ""
        subject = main_part.strip()

    return {
        "time_start": time_start,
        "time_end": "",
        "subject": subject,
        "lesson_type": "экзамен",
        "teacher": teacher,
        "room": room
    }


class SessionParser:
    def __init__(self):
        self.lessons = []
        self.subgroups = set()
        self.teachers = set()
        self.locations = set()
        self.timetables = set()

    def parse_lessons_from_bytes(self, file_bytes: bytes):
        wb = load_workbook(BytesIO(file_bytes), data_only=True)
        ws = wb.active

        for row in range(4, ws.max_row + 1):
            weekday = ws[f"A{row}"].value
            date = ws[f"B{row}"].value
            raw = ws[f"C{row}"].value

            if not weekday or not raw:
                continue

            day = WEEKDAY_MAP.get(str(weekday), 0)
            date = parse_date(date)

            parsed = parse_session_line(str(raw))

            self.lessons.append({
                "day": day,
                "date": date,
                **parsed
            })

    @property
    def teacher_location_assignments(self):
        return []

    @property
    def subjects(self):
        return set(l["subject"] for l in self.lessons)